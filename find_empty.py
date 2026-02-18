#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

excel_file = "خطة المقررات الدراسية 1447هـ.xlsx"
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

print("الأسطر التي لا تحتوي على مادة في العمود الثالث:\n")

for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is not None and row[0] != "المرحلة":
        if row[2] is None or row[2] == "":
            print(f"{row_num}: {row}")

print("\n\nالأسطر التي تحتوي على قيمة في العمود الثالث لكن ربما عمود آخر:\n")
count = 0
for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is not None and row[0] != "المرحلة" and (row[2] is None or row[2] == ""):
        # تحقق من وجود أي قيمة في الأعمدة الأخرى
        if any(val is not None for val in row):
            print(f"{row_num}: Phase={row[0]}, Grade={row[1]}, Subject={row[2]}, Periods={row[3]}, Rest={row[4:]}")
            count += 1

print(f"\nالعدد الكلي للأسطر الفارغة: {count}")
