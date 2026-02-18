#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

excel_file = "خطة المقررات الدراسية 1447هـ.xlsx"
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

# قائمة الأسطر المقروءة
read_count = 0
read_subjects = []

for row in worksheet.iter_rows(min_row=2, values_only=True):
    if row[0] is not None and row[0] != "المرحلة":
        phase = str(row[0]).strip() if row[0] else None
        grade = str(row[1]).strip() if row[1] else None
        subject = str(row[2]).strip() if row[2] else None
        periods = int(row[3]) if row[3] and isinstance(row[3], int) else 0
        
        if phase and grade and subject:
            read_count += 1
            read_subjects.append(f"{phase}|{grade}|{subject}")

print(f"عدد الأسطر المقروءة: {read_count}")
print(f"عدد الأسطر الإجمالي: 313")
print(f"الفرق: {313 - read_count}")

# الآن فحص الأسطر الكاملة
all_count = 0
all_subjects = []

for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is not None and row[0] != "المرحلة":
        all_count += 1
        key = f"{row[0]}|{row[1]}|{row[2]}"
        all_subjects.append((row_num, key))

print(f"\nعدد الأسطر مع قيم الأول والثاني والثالث: {all_count}")

# أوجد الفرق
all_keys = set(x[1] for x in all_subjects)
read_keys = set(read_subjects)

missing = all_keys - read_keys

if missing:
    print(f"\nالمواد المفقودة ({len(missing)}):")
    for key in missing:
        print(f"  {key}")
else:
    print("\nلا توجد مواد مفقودة!")
