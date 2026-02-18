import openpyxl
import json

file_path = r"خطة المقررات الدراسية 1447هـ.xlsx"
wb = openpyxl.load_workbook(file_path)

# اطبع أسماء الأوراق
print("Sheet names:")
for i, sheet in enumerate(wb.sheetnames):
    print(f"{i}: {sheet}")

# احصل على عدد الصفوف والأعمدة في كل ورقة
print("\n\nSheet Details:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n{sheet}: {ws.max_row} rows, {ws.max_column} columns")
    # اطبع أول 5 صفوف
    print("First 5 rows:")
    for i, row in enumerate(list(ws.iter_rows(min_row=1, max_row=5, values_only=True))):
        print(f"  Row {i+1}: {row[:8]}")  # أول 8 أعمدة فقط

# اختبر البحث عن الصفوف المحددة
print("\n\n=== البحث عن الصفوف المحددة ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{sheet_name}:")
    
    # ابحث عن الصفوف التي تحتوي على "الثالث" أو "الثاني"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_str = str(row)
        if "الثالث" in row_str or "الثاني" in row_str:
            print(f"  {row[:5]}")
