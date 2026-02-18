#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

# فتح ملف Excel
excel_file = "خطة المقررات الدراسية 1447هـ.xlsx"
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

# حساب عدد الأسطر بأسلوب مختلف
count = 0
print("عدد الأسطر مع القيم:\n")
for row_num, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
    if row[0] is not None:  # إذا كان هناك قيمة في العمود الأول
        if row_num > 1:  # تخطي الرأس
            count += 1
            if count <= 15:  # اعرض أول 15 أسطر
                print(f"{row_num}. {row[0][:20] if row[0] else 'None':20} | {row[1][:20] if row[1] else 'None':20} | {row[2][:30] if row[2] else 'None':30}")

print(f"\n\nإجمالي الأسطر (بدون الرأس): {count}")

# عد عدد المواد المختلفة
total = 0
for row in worksheet.iter_rows(min_row=2, values_only=True):
    if row[0] is not None and row[0] != "المرحلة" and row[2] is not None:
        total += 1

print(f"إجمالي الأسطر التي تحتوي على مادة: {total}")
