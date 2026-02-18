#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

excel_file = "خطة المقررات الدراسية 1447هـ.xlsx"
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

print("الصفوف الأولى من ملف Excel:\n")
for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    print(f"{row_num}. {row}")

print("\n\nجميع أسماء الصفوف الفريدة:")
grade_names = set()
for row in worksheet.iter_rows(min_row=2, values_only=True):
    if row[2] is not None and row[2] != "المادة الدراسية":
        grade_names.add(str(row[2]).strip())

for i, grade in enumerate(sorted(grade_names), 1):
    print(f"{i}. '{grade}'")
