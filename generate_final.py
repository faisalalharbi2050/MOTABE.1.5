# -*- coding: utf-8 -*-
import openpyxl
import re

file_path = "خطة المقررات الدراسية 1447هـ.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

# Dictionary to collect all data
data_by_key = {}

# Extract data
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    if not row[0] or not row[2]:
        continue
    
    phase = str(row[0]).strip() if row[0] else ""
    division = str(row[1]).strip() if row[1] else ""
    subject = str(row[2]).strip() if row[2] else ""
    periods = row[3] if row[3] else 0
    
    key = f"{phase}|{division}"
    
    if key not in data_by_key:
        data_by_key[key] = []
    
    data_by_key[key].append({
        'name': subject,
        'periods': int(periods) if isinstance(periods, (int, float)) else 0
    })

# Generate TypeScript code
ts_lines = ['import { Phase, Subject } from \'./types\';\n']

# ID counter for unique IDs
id_counter = 0

# Helper to create safe variable names
def safe_var_name(phase, division):
    # Create initials
    phase_abbr = phase.replace("الابتدائية", "elem").replace("المتوسطة", "mid").replace("الثانوية", "high")
    phase_abbr = phase_abbr.replace("السنة الأولى", "y1").replace("السنة الثانية", "y2").replace("السنة الثالثة", "y3")
    
    div_abbr = division.replace(" ", "_").replace("/", "_").lower()
    div_abbr = re.sub(r'[^\w_]', '', div_abbr)[:20]
    
    return f"excel_{phase_abbr}_{div_abbr}"[:50]

# Generate code for each key
for key_idx, (key, subjects) in enumerate(sorted(data_by_key.items())):
    phase, division = key.split('|')
    var_name = safe_var_name(phase, division)
    
    # Determine phase
    if "ابتدائ" in phase:
        phase_type = "Phase.ELEMENTARY"
    elif "متوسط" in phase:
        phase_type = "Phase.MIDDLE"
    else:
        phase_type = "Phase.HIGH"
    
    # Determine department
    dept = "عام"
    if "تحفيظ" in phase or "تحفيظ" in division:
        dept = "تحفيظ"
    elif "إدارة أعمال" in phase or "إدارة أعمال" in division:
        dept = "إدارة أعمال"
    elif "حاسب وهندسة" in phase or "حاسب" in division:
        dept = "حاسب وهندسة"
    elif "صحة وحياة" in phase or "صحة" in division:
        dept = "صحة وحياة"
    elif "شرعي" in phase or "شرعي" in division:
        dept = "شرعي"
    
    # Add comment
    ts_lines.append(f'\n// {phase} - {division}\nexport const {var_name}: Subject[] = [')
    
    # Add subjects
    for i, subj in enumerate(subjects):
        id_counter += 1
        subject_id = f"s_{id_counter:04d}"
        
        ts_lines.append(f'''  {{
    id: "{subject_id}",
    name: "{subj['name']}",
    specializationIds: [],
    periodsPerClass: {subj['periods']},
    phases: [{phase_type}],
    department: "{dept}",
  }},''')
    
    ts_lines.append('];\n')

# Print summary
output_str = '\n'.join(ts_lines)
print(f"عدد الصفوف من Excel: {len(data_by_key)}")
print(f"عدد المواد الكلي: {id_counter}")
print(f"عدد الأسطر: {len(ts_lines)}")

# Save to file
with open('generated_templates_final.ts', 'w', encoding='utf-8') as f:
    f.write(output_str)

print("\n✓ تم إنشاء generated_templates_final.ts")
