# -*- coding: utf-8 -*-
import openpyxl
import json

file_path = r"خطة المقررات الدراسية 1447هـ.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# احصل على أسماء الأوراق
print("Sheet names:", wb.sheetnames)

# احصل على البيانات من الورقة الأولى
ws = wb.active
print(f"\nTotal rows: {ws.max_row}, Total columns: {ws.max_column}")

# استخرج البيانات بحثاً عن الصفوف والمواد
data_by_grade = {}
current_phase = None
current_grade = None
current_division = None

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    if row[0] is None:
        continue
    
    phase = str(row[0]) if row[0] else ""
    division = str(row[1]) if row[1] else ""
    subject_name = str(row[2]) if row[2] else ""
    periods = row[3] if row[3] else 0
    
    # Print only rows with الثاني or الثالث
    if "الثاني" in subject_name or "الثالث" in subject_name or "الثاني" in phase or "الثالث" in phase:
        print(f"Row {row_idx}: Phase={phase[:20]}... | Division={division[:20]}... | Subject={subject_name[:30]}... | Periods={periods}")

# Also count subjects for specific grades
print("\n\n=== Counting subjects per grade ===")
grade_subjects = {}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None or row[2] is None:
        continue
    
    phase = str(row[0])
    division = str(row[1]) if row[1] else ""
    subject = str(row[2])
    
    # Create key for this combination
    key = f"{phase} | {division}"
    if key not in grade_subjects:
        grade_subjects[key] = 0
    grade_subjects[key] += 1

# Print results for relevant grades
for key in sorted(grade_subjects.keys()):
    if "الثالث" in key or "الثاني" in key:
        count = grade_subjects[key]
        print(f"{key}: {count} subjects")
