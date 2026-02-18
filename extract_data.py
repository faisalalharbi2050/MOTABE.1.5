# -*- coding: utf-8 -*-
import openpyxl
import json

file_path = "خطة المقررات الدراسية 1447هـ.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print("=" * 100)
print("استخراج البيانات من ملف Excel")
print("=" * 100)

# تجميع البيانات حسب المرحلة والصف والقسم
data_structure = {}

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    if not row[0] or not row[2]:  # تخطي الصفوف الفارغة
        continue
    
    phase = str(row[0]).strip() if row[0] else ""
    division = str(row[1]).strip() if row[1] else ""
    subject = str(row[2]).strip() if row[2] else ""
    periods = row[3] if row[3] else 0
    
    # إنشء مفتاح الدمج
    key = f"{phase}|{division}"
    
    if key not in data_structure:
        data_structure[key] = []
    
    data_structure[key].append({
        'subject': subject,
        'periods': periods
    })

# طباعة النتائج
for key in sorted(data_structure.keys()):
    phase, division = key.split('|')
    subjects = data_structure[key]
    
    # اطبع فقط البيانات ذات الصلة
    if "الابتدائية" in phase or "المتوسطة" in phase or "الثانوية" in phase:
        print(f"\n{'='*80}")
        print(f"المرحلة: {phase}")
        print(f"القسم/المسار: {division}")
        print(f"عدد المواد: {len(subjects)}")
        print(f"{'='*80}")
        for i, subj in enumerate(subjects, 1):
            print(f"{i:2d}. {subj['subject']:40s} ({subj['periods']} حصص)")

print("\n\n" + "=" * 100)
print("ملخص إحصائي")
print("=" * 100)
for key in sorted(data_structure.keys()):
    phase, division = key.split('|')
    if "الابتدائية" in phase or "المتوسطة" in phase or "الثانوية" in phase:
        print(f"{phase:30s} | {division:40s} : {len(data_structure[key]):3d} مواد")
